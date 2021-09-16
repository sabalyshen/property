import os
import shutil #移動檔案用

#讀取檔案
from openpyxl import load_workbook
wb = load_workbook(filename = '物品0910.xlsx', data_only=True)
ws = wb['Sheet1']

from openpyxl import load_workbook
wb2 = load_workbook(filename = '物品work.xlsx', data_only=True)
ws2 = wb2['Sheet1']

from openpyxl import load_workbook
wb3 = load_workbook(filename = '財產0910.xlsx', data_only=True)
ws3 = wb3['Sheet1']

from openpyxl import load_workbook
wb4 = load_workbook(filename = '財產work.xlsx', data_only=True)
ws4 = wb4['Sheet1']

name = []
with open('name.txt', 'r', encoding = 'utf8') as f:
    for n in f:
        n = n.replace('\n', '')
        name.append(n)
        print(n)

def find_item(list):
    for l in list:
        wb2.save(filename = l +'物品.xlsx')
        from openpyxl import load_workbook
        wb5 = load_workbook(filename =  l +'物品.xlsx', data_only=True)
        ws5 = wb5['Sheet1']
        colL = ws['L']
        a = 0
        for i in colL:
            if i.value == l:
                a += 1
                for n in range(15):
                    ws5.cell(row = 5 + a, column = 1 + n).value = ws.cell(row=i.row, column=1 + n).value
                    ws5.row_dimensions[5 + a].height = 55.8
            else:
                continue
        wb5.save(filename = l +'物品.xlsx')
               
def find_property(list):
    for l in list:
        wb4.save(filename = l +'財產.xlsx')
        from openpyxl import load_workbook
        wb6 = load_workbook(filename =  l +'財產.xlsx', data_only=True)
        ws6 = wb6['Sheet1']
        colL = ws3['L']
        a = 0
        for i in colL:
            if i.value == l:
                a += 1
                for n in range(15):
                    ws6.cell(row = 5 + a, column=1 + n).value = ws3.cell(row=i.row, column=1 + n).value
                    ws6.row_dimensions[5 + a].height = 55.8
            else:
                continue
        wb6.save(filename = l +'財產.xlsx')

while True:
    input_item = input('列印財產或是物品: (財產 = 1， 物品 = 2)')
    if input_item == '2':
        i = name
        find_item(i)
    elif input_item == '1':
        p = name
        find_property(p)
    else:
        break

source = r'C:\Users\sabal\Desktop\python'
destination = r'C:\Users\sabal\Desktop\python\物品'
files = os.listdir(source)        
for file in os.listdir('.'): # '.'代表現在資料夾的位子
    if file.endswith('物品.xlsx'):
        new_path = shutil.move(f"{source}/{file}", destination)
    elif file.endswith('財產.xlsx'):
        new_path = shutil.move(f"{source}/{file}", destination) 